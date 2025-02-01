import os
import pandas as pd
import numpy as np
from sklearn.ensemble import RandomForestClassifier
from sklearn.multioutput import MultiOutputClassifier
from sklearn.metrics import classification_report, accuracy_score
import joblib
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font

# Data directories
data_dir_train = './Training_datasets'
data_dir_test = './Test_datasets'

# List of training and test files
training_files = [f for f in os.listdir(data_dir_train) if f.endswith('.csv')]
test_files = [f for f in os.listdir(data_dir_test) if f.endswith('.csv')]


# Function to extract labels from file names
def extract_labels(file_name):
    parts = file_name.lower().split('_')
    weight = parts[0]  # "light", "medium", or "heavy"
    height = 'not_floor' if 'kneeheight' in parts else 'floor'
    return weight, height


# Function to calculate the angle between three points
def calculate_angle(x1, y1, z1, x2, y2, z2, x3, y3, z3):
    # Calculate vectors from points 2 -> 1 and 2 -> 3 (Knee -> Hip, Knee -> Ankle)
    v1 = np.array([x1 - x2, y1 - y2, z1 - z2])  # Vector from knee to hip
    v2 = np.array([x3 - x2, y3 - y2, z3 - z2])  # Vector from knee to ankle

    # Compute the cosine of the angle between the vectors
    dot_product = np.dot(v1, v2)
    magnitude_v1 = np.linalg.norm(v1)
    magnitude_v2 = np.linalg.norm(v2)
    cos_angle = dot_product / (magnitude_v1 * magnitude_v2)

    # Ensure the cosine value is in the valid range for arccos
    angle = np.arccos(np.clip(cos_angle, -1.0, 1.0))

    return np.degrees(angle)  # Convert from radians to degrees


# Function to compute summary features (mean, std, min, max)
def compute_features(data, data_for_angle):
    features = {}

    # List of body parts to extract (I only use one side assuming the other is almost the same)
    body_parts = [
        'Skeleton_25_marker:WaistLBack',  # Hip
        'SkeletonConLow:LKNE',  # Knee
        'SkeletonConLow:LANK'  # Ankle
    ]
    # Extract the X, Y, Z coordinates for hip, knee, and ankle from the data
    # Assuming 'data' contains these columns and these names are consistent.
    # Filter columns based on body parts and 'Position'
    selected_columns = data_for_angle.loc[:,
                       (data_for_angle.columns.get_level_values(0).isin(body_parts)) &
                       (data_for_angle.columns.get_level_values(1) == 'Position')
                       ]

    # Extract the time column
    time = data_for_angle.iloc[3:, 1].reset_index(drop=True).astype(float)
    # Select columns containing the relevant strings for hip, knee, and ankle positions
    x1 = selected_columns.loc[:, ('Skeleton_25_marker:WaistLBack', 'Position', 'X')].iloc[3:].reset_index(drop=True).astype(float)
    y1 = selected_columns.loc[:, ('Skeleton_25_marker:WaistLBack', 'Position', 'Y')].iloc[3:].reset_index(drop=True).astype(float)
    z1 = selected_columns.loc[:, ('Skeleton_25_marker:WaistLBack', 'Position', 'Z')].iloc[3:].reset_index(drop=True).astype(float)

    x2 = selected_columns.loc[:, ('SkeletonConLow:LKNE', 'Position', 'X')].iloc[3:].reset_index(drop=True).astype(float)
    y2 = selected_columns.loc[:, ('SkeletonConLow:LKNE', 'Position', 'Y')].iloc[3:].reset_index(drop=True).astype(float)
    z2 = selected_columns.loc[:, ('SkeletonConLow:LKNE', 'Position', 'Z')].iloc[3:].reset_index(drop=True).astype(float)

    x3 = selected_columns.loc[:, ('SkeletonConLow:LANK', 'Position', 'X')].iloc[3:].reset_index(drop=True).astype(float)
    y3 = selected_columns.loc[:, ('SkeletonConLow:LANK', 'Position', 'Y')].iloc[3:].reset_index(drop=True).astype(float)
    z3 = selected_columns.loc[:, ('SkeletonConLow:LANK', 'Position', 'Z')].iloc[3:].reset_index(drop=True).astype(float)




    # Calculate the angle for each frame (time step)
    angles = [calculate_angle(x1[i], y1[i], z1[i], x2[i], y2[i], z2[i], x3[i], y3[i], z3[i]) for i in range(len(time))]

    # Compute summary statistics for angles
    features['angle_mean'] = np.mean(angles)
    features['angle_std'] = np.std(angles)
    features['angle_min'] = np.min(angles)
    features['angle_max'] = np.max(angles)

    for col in data.columns[2:]:
        if data[col].dtype in ['float64', 'int64']:  # Only numerical columns
            features[f'{col}_mean'] = data[col].mean()
            features[f'{col}_std'] = data[col].std()
            features[f'{col}_min'] = data[col].min()
            features[f'{col}_max'] = data[col].max()
    return features


# Helper function to load and prepare data from files
def prepare_data(file_list, base_dir):
    all_data = []
    labels_weight = []
    labels_height = []
    for file_name in file_list:
        file_path = os.path.join(base_dir, file_name)
        data = pd.read_csv(file_path, skiprows=2)  # Read raw data
        data_for_angle = pd.read_csv(file_path, header=[0, 1, 2])
        weight, height = extract_labels(file_name)  # Extract labels
        features = compute_features(data, data_for_angle)  # Compute summary statistics
        all_data.append(features)  # Store computed features
        labels_weight.append(weight)  # Store weight label
        labels_height.append(height)  # Store height label

    # Convert the list of feature dictionaries into a DataFrame
    features_df = pd.DataFrame(all_data)
    return features_df, labels_weight, labels_height


# Prepare training and test datasets
X_train, y_train_weight, y_train_height = prepare_data(training_files, data_dir_train)
X_test, y_test_weight, y_test_height = prepare_data(test_files, data_dir_test)

# Ensure all data is numerical (fill missing values with 0)
X_train = X_train.fillna(0)
X_test = X_test.fillna(0)

# Label encoding for training and test labels (weight and height)
y_train = pd.DataFrame({
    'weight': pd.Categorical(y_train_weight).codes,
    'height': pd.Categorical(y_train_height).codes
})
y_test = pd.DataFrame({
    'weight': pd.Categorical(y_test_weight).codes,
    'height': pd.Categorical(y_test_height).codes
})

# Get the label categories for decoding predictions later
weight_map = pd.Categorical(y_train_weight).categories
height_map = pd.Categorical(y_train_height).categories

# Train the model
multi_output_model = MultiOutputClassifier(RandomForestClassifier(random_state=42))
multi_output_model.fit(X_train, y_train)

# Predictions on the test data
y_pred = multi_output_model.predict(X_test)
y_pred_df = pd.DataFrame(y_pred, columns=y_test.columns)

# Decode the predicted labels back to their original string values
y_pred_decoded = {
    'weight': [weight_map[pred] for pred in y_pred_df['weight']],
    'height': [height_map[pred] for pred in y_pred_df['height']]
}

# Print predicted labels and actual labels for the test files
print("Predicted vs Actual Labels for Test Files:")
for i, file_name in enumerate(test_files):
    weight_pred = y_pred_decoded['weight'][i]
    height_pred = y_pred_decoded['height'][i]
    weight_actual = y_test_weight[i]
    height_actual = y_test_height[i]
    print(
        f"File: {file_name}, Predicted Weight: {weight_pred}, Actual Weight: {weight_actual}, Predicted Height: {height_pred}, Actual Height: {height_actual}")

# Evaluation of the model
print("\nWeight Classification Report:")
print(classification_report(y_test['weight'], y_pred_df['weight']))

print("Height Classification Report:")
print(classification_report(y_test['height'], y_pred_df['height']))

# Accuracy for each label
weight_accuracy = accuracy_score(y_test['weight'], y_pred_df['weight'])
height_accuracy = accuracy_score(y_test['height'], y_pred_df['height'])
print("Weight Classification Accuracy:", weight_accuracy)
print("Height Classification Accuracy:", height_accuracy)

# Calculate the accuracy of getting the correct weight and height pair
correct_pairs = 0
total_pairs = len(test_files)

# Loop through the predictions and check if both weight and height are correct for each test file
for i in range(total_pairs):
    if y_pred_decoded['weight'][i] == y_test_weight[i] and y_pred_decoded['height'][i] == y_test_height[i]:
        correct_pairs += 1

# Calculate the accuracy of getting both weight and height correct together
pair_accuracy = correct_pairs / total_pairs

# Print the pair accuracy
print(f"\nPair Accuracy of Correct Weight and Height Predictions: {pair_accuracy:.2f}")

# Create a DataFrame to store predicted and actual values for each file, and correctness columns
results = {
    'File': test_files,
    'Predicted Weight': y_pred_decoded['weight'],
    'Actual Weight': y_test_weight,
    'Weight Correct': [pred == actual for pred, actual in zip(y_pred_decoded['weight'], y_test_weight)],
    'Predicted Height': y_pred_decoded['height'],
    'Actual Height': y_test_height,
    'Height Correct': [pred == actual for pred, actual in zip(y_pred_decoded['height'], y_test_height)]
}

# Convert the dictionary into a pandas DataFrame
results_df = pd.DataFrame(results)

# Save test results to excel file
wb = Workbook()
ws = wb.active
ws.title = "Predictions"

# Write the column headers
for col_num, column in enumerate(results_df.columns, 1):
    cell = ws.cell(row=1, column=col_num, value=column)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Write the data rows
for row_num, row in enumerate(results_df.itertuples(), 2):
    for col_num, value in enumerate(row[1:], 1):
        cell = ws.cell(row=row_num, column=col_num, value=value)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # green for correct, red for incorrect
        if col_num == 4 and value:  # "Weight Correct" column
            cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
        elif col_num == 4 and not value:  # Incorrect weight prediction
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
        if col_num == 7 and value:  # "Height Correct" column
            cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
        elif col_num == 7 and not value:  # Incorrect height prediction
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red

# Adjust the width of the "File" column based on the largest file name length
file_column_width = max(results_df['File'].apply(lambda x: len(str(x)))) + 2
ws.column_dimensions['A'].width = file_column_width

# Adjust the width of other columns
for col_num in range(2, len(results_df.columns) + 1):
    column_letter = chr(64 + col_num)
    ws.column_dimensions[column_letter].width = file_column_width * 1 / 2

# Save the test results to excel file
output_file = 'testset-results.xlsx'
wb.save(output_file)

# Save the model to pickle file
joblib.dump(multi_output_model, 'height-weight-classification-model.pkl')
print("Multi-output model saved as 'height-weight-classification-model.pkl'")
